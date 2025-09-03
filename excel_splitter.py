import os
import sys
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox, ttk
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
import threading
import queue

# --- Global Style Cache and Function Placeholders ---
style_cache = {}
_copy_cell_properties = None
_copy_row_formatting = None

# --- UI and Core Logic Classes ---

class ProgressManager:
    """Manages the Toplevel progress window UI."""
    def __init__(self, parent, title, total_steps, cancel_event):
        self.parent = parent
        self.total_steps = total_steps
        self.cancel_event = cancel_event
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.request_cancel)
        self.status_label = tk.Label(self.window, text="Initializing...", padx=20, pady=10, width=60)
        self.status_label.pack()
        self.progress_bar = ttk.Progressbar(self.window, orient="horizontal", length=400, mode="determinate", maximum=total_steps)
        self.progress_bar.pack(padx=20, pady=5)
        self.cancel_button = tk.Button(self.window, text="Cancel", command=self.request_cancel, width=10)
        self.cancel_button.pack(pady=10)
        self.parent.update_idletasks()
        
    def update(self, current_step, status_text):
        """Updates the GUI and terminal progress indicators."""
        self.progress_bar['value'] = current_step
        self.status_label.config(text=status_text)
        progress_percent = (current_step / self.total_steps) * 100
        bar_length = 30
        filled_length = int(bar_length * current_step // self.total_steps)
        bar = '█' * filled_length + '-' * (bar_length - filled_length)
        terminal_text = f"\rProgress: |{bar}| {progress_percent:.1f}% ({current_step}/{self.total_steps}) - Processing..."
        sys.stdout.write(terminal_text)
        sys.stdout.flush()
        self.parent.update_idletasks()

    def request_cancel(self):
        """Flags the operation for cancellation after user confirmation."""
        if messagebox.askyesno("Confirm Cancel", "Are you sure you want to cancel the operation?"):
            self.status_label.config(text="Cancellation requested...")
            self.cancel_event.set()

    def close(self):
        """Closes the progress window."""
        sys.stdout.write('\n')
        self.window.destroy()

# --- Copying Functions Version 1 (Standard) ---
def _copy_cell_properties_v1(source_cell, target_cell):
    """Standard (slow) method: Copies cell value and all style attributes individually."""
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.hyperlink:
        target_cell.hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)

def _copy_row_formatting_v1(ws_source, ws_target, source_row_idx, target_row_idx, max_col):
    """Standard (slow) method: Copies an entire row by iterating through each cell."""
    for col_idx in range(1, max_col + 1):
        source_cell = ws_source.cell(row=source_row_idx, column=col_idx)
        target_cell = ws_target.cell(row=target_row_idx, column=col_idx)
        _copy_cell_properties(source_cell, target_cell)
    if source_row_idx in ws_source.row_dimensions:
        ws_target.row_dimensions[target_row_idx].height = ws_source.row_dimensions[source_row_idx].height

# --- Copying Functions Version 2 (Optimized) ---
def _copy_cell_properties_v2(source_cell, target_cell):
    """Optimized (fast) method: Copies cell value and reuses style objects via a cache."""
    target_cell.value = source_cell.value
    if source_cell.has_style:
        style_key = source_cell._style
        if style_key in style_cache:
            target_cell._style = style_cache[style_key]
        else:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
            style_cache[style_key] = target_cell._style
    if source_cell.hyperlink:
        target_cell.hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)

def _copy_row_formatting_v2(ws_source, ws_target, source_row_idx, target_row_idx, max_col):
    """Optimized (fast) method: Copies an entire row by iterating through each cell."""
    for col_idx in range(1, max_col + 1):
        source_cell = ws_source.cell(row=source_row_idx, column=col_idx)
        target_cell = ws_target.cell(row=target_row_idx, column=col_idx)
        _copy_cell_properties(source_cell, target_cell)
    if source_row_idx in ws_source.row_dimensions:
        ws_target.row_dimensions[target_row_idx].height = ws_source.row_dimensions[source_row_idx].height

# --- Other Helper Functions ---
def _copy_merged_cells(ws_source, ws_target, min_row, max_row, row_offset):
    """Copies and adjusts merged cell ranges from a source to a target worksheet."""
    for merged_range in ws_source.merged_cells.ranges:
        if merged_range.min_row >= min_row and merged_range.max_row <= max_row:
            new_min_row = merged_range.min_row - row_offset
            new_max_row = merged_range.max_row - row_offset
            try:
                ws_target.merge_cells(start_row=new_min_row, start_column=merged_range.min_col,
                                      end_row=new_max_row, end_column=merged_range.max_col)
            except Exception as e:
                print(f"Warning: Could not merge range {merged_range}: {e}")

# --- Worker Function ---
def split_excel_file_with_formatting(input_file, output_directory, chunk_size, heading_rows, preserve_formulas, progress_queue, cancel_event):
    """Performs the Excel splitting in a worker thread."""
    input_filename_base = os.path.splitext(os.path.basename(input_file))[0]
    files_created = 0
    
    try:
        wb_source = openpyxl.load_workbook(input_file, data_only=not preserve_formulas)
        ws_source = wb_source.active
    except Exception as e:
        progress_queue.put({'type': 'result', 'data': {'status': 'error', 'message': f"Error loading Excel file: {e}", 'files_created': 0}})
        return

    total_rows, max_col = ws_source.max_row, ws_source.max_column
    if total_rows == 0:
        progress_queue.put({'type': 'result', 'data': {'status': 'success', 'message': "Input file's active sheet was empty.", 'files_created': 0}})
        return
    
    data_rows_to_process = total_rows - heading_rows
    if data_rows_to_process <= 0:
        progress_queue.put({'type': 'result', 'data': {'status': 'error', 'message': 'No data rows to process after accounting for header rows.', 'files_created': 0}})
        return
    
    num_chunks = (data_rows_to_process + chunk_size - 1) // chunk_size

    for i in range(num_chunks):
        if cancel_event.is_set():
            progress_queue.put({'type': 'result', 'data': {'status': 'cancelled', 'message': 'Operation cancelled.', 'files_created': files_created}})
            return

        source_data_start_row = heading_rows + (i * chunk_size) + 1
        source_data_end_row = min(heading_rows + ((i + 1) * chunk_size), total_rows)
        
        status_text = f"Processing chunk {i+1}/{num_chunks}..."
        progress_queue.put({'type': 'progress', 'step': i + 1, 'status': status_text})
        
        style_cache.clear()
        wb_chunk = openpyxl.Workbook()
        ws_chunk = wb_chunk.active
        ws_chunk.title = ws_source.title

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in ws_source.column_dimensions:
                ws_chunk.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width

        current_target_row = 1
        if heading_rows > 0:
            for r_idx in range(1, heading_rows + 1):
                _copy_row_formatting(ws_source, ws_chunk, r_idx, current_target_row, max_col)
                current_target_row += 1
            _copy_merged_cells(ws_source, ws_chunk, 1, heading_rows, 0)

        data_row_offset = source_data_start_row - current_target_row
        for source_r_idx in range(source_data_start_row, source_data_end_row + 1):
            _copy_row_formatting(ws_source, ws_chunk, source_r_idx, current_target_row, max_col)
            current_target_row += 1
        _copy_merged_cells(ws_source, ws_chunk, source_data_start_row, source_data_end_row, data_row_offset)

        # --- FIX: Reverted to the original, correct filename format ---
        output_file_name = f"{input_filename_base}_rows_{source_data_start_row}-{source_data_end_row}.xlsx"
        output_path = os.path.join(output_directory, output_file_name)
        try:
            wb_chunk.save(output_path)
            files_created += 1
        except Exception as e:
            progress_queue.put({'type': 'result', 'data': {'status': 'error', 'message': f"Error saving {output_path}: {e}", 'files_created': files_created}})
            return
    
    progress_queue.put({'type': 'result', 'data': {'status': 'success', 'message': f'Successfully created {files_created} files.', 'files_created': files_created}})

# --- Main Application Class ---
class App:
    """The main application class that orchestrates the GUI and the worker thread."""
    def __init__(self, root):
        self.root = root
        self.progress_manager = None
        self.run()

    def get_user_input(self):
        """Handles all initial user dialogs to get processing parameters."""
        self.input_file = filedialog.askopenfilename(title="Select Input Excel File (.xlsx)", filetypes=[("Excel files", "*.xlsx")])
        if not self.input_file: return False
        
        self.output_directory = filedialog.askdirectory(title="Select Output Directory", initialdir=os.path.dirname(self.input_file))
        if not self.output_directory: return False

        use_version2 = messagebox.askyesno(
            title="Select Cell Copying Method",
            message="Please choose the method for copying cell styles.\n\n"
                    "• Optimized (Recommended): Uses style caching for a significant speed boost.\n\n"
                    "• Standard: Slower, direct-copy method for baseline compatibility.\n\n"
                    "Do you want to use the Optimized method?"
        )
        global _copy_cell_properties, _copy_row_formatting
        if use_version2:
            _copy_cell_properties = _copy_cell_properties_v2
            _copy_row_formatting = _copy_row_formatting_v2
            print("Using optimized copy functions (Style Caching).")
        else:
            _copy_cell_properties = _copy_cell_properties_v1
            _copy_row_formatting = _copy_row_formatting_v1
            print("Using standard copy functions.")

        try:
            self.chunk_size = simpledialog.askinteger("Data Rows Per File", "Enter data rows per file (excluding headers):", initialvalue=50000, minvalue=1)
            if self.chunk_size is None: return False
            self.heading_rows = simpledialog.askinteger("Header Rows", "Enter header rows to repeat in each file:", initialvalue=1, minvalue=0)
            if self.heading_rows is None: return False
        except (TypeError, ValueError): return False

        self.preserve_formulas = messagebox.askyesno(
            title="Preserve Formulas?",
            message="Do you want to preserve formulas?\n\n"
                    "• 'Yes' will keep formulas (e.g., '=A1+B1'), but may cause #REF! errors.\n"
                    "• 'No' will copy only calculated values (e.g., '123'), ensuring data is static."
        )
        return True

    def start_processing(self):
        """Prepares for and launches the background processing thread."""
        print("\n--- Settings ---")
        print(f"  Input file: {self.input_file}")
        print(f"  Output directory: {self.output_directory}")
        print(f"  Chunk size: {self.chunk_size}")
        print(f"  Header rows: {self.heading_rows}")
        print(f"  Preserve Formulas: {'Yes' if self.preserve_formulas else 'No'}")
        print("------------------\n")

        try:
            wb_check = openpyxl.load_workbook(self.input_file, read_only=True)
            total_rows_check = wb_check.active.max_row
            wb_check.close()
            data_rows_check = total_rows_check - self.heading_rows
            num_chunks_check = (data_rows_check + self.chunk_size - 1) // self.chunk_size if data_rows_check > 0 else 0
        except Exception as e:
            messagebox.showerror("Error", f"Could not read input file to determine size: {e}")
            return

        if num_chunks_check <= 0:
            messagebox.showinfo("Information", "No data rows to process based on the settings provided.")
            return

        self.progress_queue = queue.Queue()
        self.cancel_event = threading.Event()
        self.progress_manager = ProgressManager(self.root, "Splitting File...", num_chunks_check, self.cancel_event)
        
        self.worker_thread = threading.Thread(
            target=split_excel_file_with_formatting,
            args=(self.input_file, self.output_directory, self.chunk_size, self.heading_rows, self.preserve_formulas, self.progress_queue, self.cancel_event)
        )
        self.worker_thread.start()
        self.root.after(100, self.check_queue)

    def check_queue(self):
        """Periodically checks the queue for messages from the worker thread."""
        try:
            while True:
                message = self.progress_queue.get(block=False)
                if message['type'] == 'progress':
                    self.progress_manager.update(message['step'], message['status'])
                elif message['type'] == 'result':
                    self.on_task_finished(message['data'])
                    return
        except queue.Empty:
            pass

        if self.worker_thread.is_alive():
            self.root.after(100, self.check_queue)
        else:
            self.on_task_finished({'status': 'error', 'message': 'The worker thread terminated unexpectedly.', 'files_created': 'Unknown'})

    def on_task_finished(self, result):
        """Handles the final result from the worker and displays a summary."""
        if self.progress_manager:
            self.progress_manager.close()

        files_created = result.get('files_created', 'N/A')
        print("\n--- Operation Summary ---")
        print(f"Status: {result.get('status', 'Unknown').title()}")
        print(f"Message: {result.get('message', 'No message.')}")
        print(f"Total Files Created: {files_created}")
        print("-------------------------\n")
        
        status_map = {'error': messagebox.showerror, 'cancelled': messagebox.showwarning, 'success': messagebox.showinfo}
        title = result.get('status', 'Info').title()
        message = f"{result.get('message', '')}\n\nTotal files created: {files_created}"
        if result.get('status') == 'success':
            message = f"{result.get('message', '')}\n\nOutput directory:\n{self.output_directory}"

        status_map.get(result.get('status'), messagebox.showinfo)(title, message)
        self.root.destroy()

    def run(self):
        """Main execution flow of the application."""
        print("SplitXL")
        print("=" * 40)
        if self.get_user_input():
            self.start_processing()
        else:
            print("\nOperation cancelled during setup. Exiting...")
            self.root.destroy()

if __name__ == "__main__":
    try:
        root = tk.Tk()
        root.attributes('-alpha', 0.0)
        root.withdraw()
        app = App(root)
        root.mainloop()
    except tk.TclError as e:
        print(f"Failed to start GUI application: {e}")
        print("This script requires a graphical desktop environment to run.")